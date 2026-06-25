#!/usr/bin/env python3
"""
validador.py — Validação de Folha de Pagamento
Uso:
    python validador.py --pasta ./arquivos_06-2026
    python validador.py --extrato folha.pdf --apontamentos apo.xls --emprestimos emp.xlsx
    python validador.py --extrato folha.pdf --apontamentos apo.xls --emprestimos emp.xlsx --coop desc_coop.xls
"""

import re, io, math, sys, os, argparse, datetime, glob
from dataclasses import dataclass, field
from typing import Optional

# ── Dependências ──────────────────────────────────────────────────────────────
try:
    import pdfplumber
except ImportError:
    sys.exit("❌  Instale as dependências: pip install pdfplumber pandas openpyxl xlrd")
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════════════════════
# Estruturas
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class FuncionarioFolha:
    matricula: int
    nome: str
    salario: float = 0.0
    horas_mes: float = 220.0
    rubricas: dict = field(default_factory=dict)
    emprestimos: list = field(default_factory=list)


@dataclass
class Divergencia:
    tipo: str       # APONTAMENTOS | EMPRESTIMOS | COPARTICIPACAO
    matricula: int
    nome: str
    rubrica: str
    descricao: str
    valor_excel: float
    valor_folha: float
    diff: float
    gravidade: str  # ERRO | AVISO


TOLERANCE = 0.10


# ═══════════════════════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _to_float(v) -> float:
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return 0.0
    if isinstance(v, str):
        v = v.replace('R$', '').strip()
        if re.search(r'\d\.\d{3},', v):
            v = v.replace('.', '').replace(',', '.')
        else:
            v = v.replace(',', '.')
        v = re.sub(r'[^\d.]', '', v)
        try: return float(v)
        except ValueError: return 0.0
    try: return float(v)
    except: return 0.0


def _is_cancelado(v) -> bool:
    return isinstance(v, str) and 'CANCEL' in v.upper()


def _parse_rubrica_line(line: str) -> list:
    """Extrai (cod, val, ind) de linha podendo ter 2 rubricas lado a lado."""
    results = []
    segments = re.split(r'(?<=[ PD\*])\s+(?=\d{1,5}[^\d,.])', line)
    for seg in segments:
        seg = seg.strip()
        m = re.match(r'^(\d{1,5})', seg)
        if not m: continue
        cod = m.group(1)
        mv = re.search(r'([\d]{1,3}(?:\.\d{3})*(?:,\d+)?)\s*([PD\*])\s*$', seg)
        if not mv: continue
        val = _to_float(mv.group(1))
        if val > 0:
            results.append((cod, val, mv.group(2)))
    return results


# ═══════════════════════════════════════════════════════════════════════════════
# Leitura PDF Extrato Mensal (Domínio)
# ═══════════════════════════════════════════════════════════════════════════════

_EMP_HEAD    = re.compile(r'Empr\.\:\s*(\d+)\s*(.+?)\s+Situa[çc]')
_CONTR_HEAD  = re.compile(r'Contr\:\s*(\d+)\s*(.+?)\s+Situa[çc]')
_SAL_LINE    = re.compile(r'Sal[aá]rio:\s*([\d.,]+)')
_HORAS_LINE  = re.compile(r'Horas\s+M[eê]s:\s*([\d.,]+)')
_LOAN_LINE   = re.compile(
    r'^(\d{3,4})\s+DESC[\.\s]+EMP[\.\s]+CRED[\.\s]+TRAB\s+N[oOº°]?\s*([\w]+)\s+'
    r'([\d.,]+)\s+([\d.,]+)\s*D\s*$', re.IGNORECASE)
_EMPRESA_LINE = re.compile(r'Empresa:\s*\d+\s*-\s*(.+?)(?:\s+P[áa]gina|\s+CNPJ|$)')
_COMP_LINE   = re.compile(r'Compet[êe]ncia:\s*(\d{2}/\d{4})')


def parse_extrato_pdf(path: str) -> tuple:
    """Retorna (dict{mat: FuncionarioFolha}, empresa, competencia)."""
    funcionarios: dict[int, FuncionarioFolha] = {}
    current: Optional[FuncionarioFolha] = None
    empresa = competencia = ''

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text(x_tolerance=3, y_tolerance=3) or ''
            for line in text.splitlines():
                line = line.strip()
                if not line: continue
                if not empresa:
                    m = _EMPRESA_LINE.search(line)
                    if m: empresa = m.group(1).strip()
                if not competencia:
                    m = _COMP_LINE.search(line)
                    if m: competencia = m.group(1)
                m = _EMP_HEAD.search(line) or _CONTR_HEAD.search(line)
                if m:
                    if current: funcionarios[current.matricula] = current
                    nome = re.split(r'\s+(?:CPF|Situa)', m.group(2))[0].strip()
                    current = FuncionarioFolha(matricula=int(m.group(1)), nome=nome)
                    continue
                if current is None: continue
                ms = _SAL_LINE.search(line)
                if ms and current.salario == 0.0:
                    current.salario = _to_float(ms.group(1))
                mh = _HORAS_LINE.search(line)
                if mh:
                    v = _to_float(mh.group(1))
                    if v > 0: current.horas_mes = v
                ml = _LOAN_LINE.match(line)
                if ml:
                    valor = _to_float(ml.group(4))
                    if valor > 0:
                        current.emprestimos.append({
                            'contrato': ml.group(2).strip(),
                            'valor': valor,
                            'rubrica': ml.group(1)
                        })
                    continue
                for cod, val, _ in _parse_rubrica_line(line):
                    current.rubricas[cod] = current.rubricas.get(cod, 0.0) + val

    if current: funcionarios[current.matricula] = current
    return funcionarios, empresa, competencia


# ═══════════════════════════════════════════════════════════════════════════════
# Leitura Excel Apontamentos
# ═══════════════════════════════════════════════════════════════════════════════

def parse_apontamentos(path: str) -> dict:
    ext = path.lower().split('.')[-1]
    engine = 'xlrd' if ext == 'xls' else 'openpyxl'
    raw = pd.read_excel(path, header=None, engine=engine)

    header_row = code_row = None
    for i, row in raw.iterrows():
        vals = [str(v) for v in row.values]
        if any('digo Empregado' in v or 'Codigo Empregado' in v for v in vals):
            header_row = i
        if header_row is not None and i == header_row + 1:
            code_row = i
            break

    if header_row is None:
        raise ValueError(f"Apontamentos ({path}): cabeçalho 'Código Empregado' não encontrado.")

    col_codes = {}
    if code_row is not None:
        for col_i, val in enumerate(raw.iloc[code_row]):
            s = str(val).strip().split('.')[0]
            if s.isdigit():
                col_codes[col_i] = s

    result = {}
    for idx in range(header_row + 2, len(raw)):
        row = raw.iloc[idx]
        mat_raw = row.iloc[1]
        if pd.isna(mat_raw): continue
        try: mat = int(float(str(mat_raw).strip()))
        except ValueError: continue
        if mat <= 0: continue
        result[mat] = {
            rub: (0.0 if _is_cancelado(row.iloc[ci]) else _to_float(row.iloc[ci]))
            for ci, rub in col_codes.items()
        }
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# Leitura Excel Empréstimos
# ═══════════════════════════════════════════════════════════════════════════════

def parse_emprestimos(path: str) -> list:
    ext = path.lower().split('.')[-1]
    engine = 'xlrd' if ext == 'xls' else 'openpyxl'
    df = pd.read_excel(path, engine=engine)
    df.columns = [str(c).strip() for c in df.columns]

    if 'contrato' not in df.columns or 'valorParcela' not in df.columns:
        raise ValueError(f"Empréstimos ({path}): colunas 'contrato'/'valorParcela' não encontradas.")

    records = []
    for _, row in df.iterrows():
        contrato = str(row.get('contrato', '')).strip()
        if not contrato or contrato.lower() in ('nan', 'none', ''): continue
        parcela = _to_float(row.get('valorParcela'))
        if parcela <= 0: continue
        records.append({
            'nome': str(row.get('nomeTrabalhador', '')).strip(),
            'contrato': contrato,
            'valorParcela': parcela,
        })
    return records


# ═══════════════════════════════════════════════════════════════════════════════
# Leitura Excel Coparticipação
# ═══════════════════════════════════════════════════════════════════════════════

def parse_coop_excel(path: str) -> dict:
    ext = path.lower().split('.')[-1]
    engine = 'xlrd' if ext == 'xls' else 'openpyxl'
    raw = pd.read_excel(path, header=None, engine=engine)

    header_row = None
    for i, row in raw.iterrows():
        vals = [str(v).strip() for v in row.values]
        if any(v in ('Cód.', 'Cod.', 'Código') for v in vals):
            header_row = i
            break
    if header_row is None:
        raise ValueError(f"Coparticipação ({path}): cabeçalho não reconhecido.")

    header = raw.iloc[header_row].tolist()
    col_cod = col_227 = col_203 = None
    for i, h in enumerate(header):
        h = str(h).strip()
        if h in ('Cód.', 'Cod.', 'Código'): col_cod = i
        elif '227' in h: col_227 = i
        elif '203' in h: col_203 = i
    if col_cod is None: col_cod = 1
    if col_227 is None: col_227 = 3
    if col_203 is None: col_203 = 6

    result = {}
    for idx in range(header_row + 1, len(raw)):
        row = raw.iloc[idx]
        mat_raw = row.iloc[col_cod]
        if pd.isna(mat_raw): continue
        try: mat = int(float(str(mat_raw).strip()))
        except ValueError: continue
        if mat <= 0: continue
        v227 = _to_float(row.iloc[col_227])
        v203 = _to_float(row.iloc[col_203]) if col_203 < len(row) else 0.0
        if mat in result:
            result[mat]['227'] += v227
            result[mat]['203'] += v203
        else:
            result[mat] = {'227': v227, '203': v203}
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# Validações
# ═══════════════════════════════════════════════════════════════════════════════

RUBRICAS_DIRETAS = {
    '227': 'Plano Saúde Informativa',
    '230': 'VR Local Informativa',
    '231': 'VR Cartão Informativa',
    '224': 'VT Cartão Informativa',
    '8111': 'Desconto Plano de Saúde',
    '202': 'Assistência Odontológica',
    '203': 'Fator Moderador Unimed',
    '204': 'Desconto Vale Refeição',
    '48':  'Desconto Vale Transporte',
}

RUBRICAS_HE = {
    '150': ('Horas Extras 50%',  1.50),
    '200': ('Horas Extras 100%', 2.00),
}


def validate_apontamentos(apontamentos: dict, folha: dict) -> list:
    divs = []
    for mat, apo in apontamentos.items():
        func = folha.get(mat)
        if func is None:
            divs.append(Divergencia(
                tipo='APONTAMENTOS', matricula=mat, nome=f'Matrícula {mat}',
                rubrica='—', descricao='Matrícula no Excel não encontrada na folha',
                valor_excel=0, valor_folha=0, diff=0, gravidade='AVISO'
            ))
            continue
        for cod, desc in RUBRICAS_DIRETAS.items():
            vx = apo.get(cod, 0.0)
            vf = func.rubricas.get(cod, 0.0)
            if abs(vx - vf) > TOLERANCE:
                divs.append(Divergencia(
                    tipo='APONTAMENTOS', matricula=mat, nome=func.nome,
                    rubrica=cod, descricao=desc,
                    valor_excel=vx, valor_folha=vf, diff=vx - vf,
                    gravidade='ERRO' if abs(vx - vf) > 0.50 else 'AVISO'
                ))
        hm = func.horas_mes if func.horas_mes > 0 else 220.0
        sal_hora = func.salario / hm if func.salario > 0 else 0.0
        for cod, (desc, mult) in RUBRICAS_HE.items():
            horas = apo.get(cod, 0.0)
            vf = func.rubricas.get(cod, 0.0)
            if horas == 0 and vf == 0: continue
            vx = round(horas * sal_hora * mult, 2) if sal_hora > 0 else 0.0
            if abs(vx - vf) > TOLERANCE:
                divs.append(Divergencia(
                    tipo='APONTAMENTOS', matricula=mat, nome=func.nome,
                    rubrica=cod,
                    descricao=f'{desc} — {horas}h × R${sal_hora:.4f}/h × {mult} = R${vx:.2f}',
                    valor_excel=vx, valor_folha=vf, diff=vx - vf, gravidade='ERRO'
                ))
    return divs


def validate_emprestimos(emprestimos_excel: list, folha: dict) -> list:
    divs = []
    folha_contratos = {}
    for func in folha.values():
        for loan in func.emprestimos:
            folha_contratos[loan['contrato']] = {
                'valor': loan['valor'], 'rubrica': loan['rubrica'],
                'matricula': func.matricula, 'nome': func.nome
            }
    excel_contratos = set()
    for rec in emprestimos_excel:
        c = rec['contrato']
        excel_contratos.add(c)
        vx = rec['valorParcela']
        if c in folha_contratos:
            vf = folha_contratos[c]['valor']
            if abs(vx - vf) > TOLERANCE:
                divs.append(Divergencia(
                    tipo='EMPRESTIMOS',
                    matricula=folha_contratos[c]['matricula'],
                    nome=folha_contratos[c]['nome'],
                    rubrica=folha_contratos[c]['rubrica'],
                    descricao=f'Parcela do contrato {c}',
                    valor_excel=vx, valor_folha=vf, diff=vx - vf, gravidade='ERRO'
                ))
        else:
            divs.append(Divergencia(
                tipo='EMPRESTIMOS', matricula=0, nome=rec['nome'],
                rubrica='—',
                descricao=f'Contrato {c} no Excel porém NÃO DESCONTADO na folha',
                valor_excel=vx, valor_folha=0.0, diff=vx, gravidade='ERRO'
            ))
    for c, info in folha_contratos.items():
        if c not in excel_contratos:
            divs.append(Divergencia(
                tipo='EMPRESTIMOS', matricula=info['matricula'], nome=info['nome'],
                rubrica=info['rubrica'],
                descricao=f'Contrato {c} descontado na folha porém AUSENTE no Excel',
                valor_excel=0.0, valor_folha=info['valor'], diff=-info['valor'], gravidade='ERRO'
            ))
    return divs


def validate_coop(coop_excel: dict, folha: dict) -> list:
    divs = []
    rubricas = {'227': 'Plano Saúde Informativa', '203': 'Fator Moderador Unimed'}
    all_mats = set(coop_excel.keys()) | {
        m for m, f in folha.items() if '227' in f.rubricas or '203' in f.rubricas
    }
    for mat in all_mats:
        func = folha.get(mat)
        nome = func.nome if func else f'Matrícula {mat}'
        ed = coop_excel.get(mat, {})
        for cod, desc in rubricas.items():
            vx = ed.get(cod, 0.0)
            vf = func.rubricas.get(cod, 0.0) if func else 0.0
            if abs(vx - vf) > TOLERANCE:
                divs.append(Divergencia(
                    tipo='COPARTICIPACAO', matricula=mat, nome=nome,
                    rubrica=cod, descricao=desc,
                    valor_excel=vx, valor_folha=vf, diff=vx - vf,
                    gravidade='ERRO' if abs(vx - vf) > 1.0 else 'AVISO'
                ))
    return divs


# ═══════════════════════════════════════════════════════════════════════════════
# Relatório Excel
# ═══════════════════════════════════════════════════════════════════════════════

def _header_row(ws, row, cols, cor='1F3864'):
    fill = PatternFill('solid', fgColor=cor)
    font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill; cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def _borda():
    s = Side(style='thin', color='D0D0D0')
    return Border(left=s, right=s, top=s, bottom=s)

def _aba_divs(ws, divs: list, titulo: str):
    ws.title = titulo[:31]
    headers = ['Matrícula','Nome','Rubrica','Descrição','Valor Excel','Valor Folha','Diferença','Gravidade']
    widths   = [10, 40, 9, 54, 13, 13, 13, 10]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _header_row(ws, 1, len(headers))
    f_err  = PatternFill('solid', fgColor='FFDCE0')
    f_warn = PatternFill('solid', fgColor='FFF3CD')
    fn = Font(name='Arial', size=9)
    fb = Font(name='Arial', size=9, bold=True)
    b = _borda()
    for r, d in enumerate(sorted(divs, key=lambda x: (x.gravidade, x.matricula)), 2):
        fill = f_err if d.gravidade == 'ERRO' else f_warn
        vals = [d.matricula or '', d.nome, d.rubrica, d.descricao,
                d.valor_excel, d.valor_folha, d.diff, d.gravidade]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.fill = fill
            cell.font = fb if d.gravidade == 'ERRO' else fn
            cell.border = b
            cell.alignment = Alignment(vertical='center')
            if col in (5, 6, 7):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
    if not divs:
        c = ws.cell(row=2, column=1, value='✅  Nenhuma divergência encontrada.')
        c.font = Font(name='Arial', size=10, color='1A7F37', bold=True)


def gerar_relatorio(divergencias: list, competencia: str, empresa: str, n_func: int, destino: str):
    wb = Workbook()
    ws_r = wb.active
    ws_r.title = 'Resumo'
    ws_r.column_dimensions['A'].width = 32
    ws_r.column_dimensions['B'].width = 24

    ws_r.merge_cells('A1:B1')
    t = ws_r['A1']
    t.value = f'Validação de Folha — {competencia}'
    t.font = Font(name='Arial', size=14, bold=True, color='1F3864')
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws_r.row_dimensions[1].height = 32

    meta = [('Empresa', empresa), ('Competência', competencia),
            ('Funcionários na folha', str(n_func)),
            ('Emissão', datetime.datetime.now().strftime('%d/%m/%Y %H:%M'))]
    for i, (k, v) in enumerate(meta, 2):
        ws_r.cell(row=i, column=1, value=k).font = Font(name='Arial', size=10, bold=True)
        ws_r.cell(row=i, column=2, value=v).font = Font(name='Arial', size=10)

    ws_r.cell(row=7, column=1, value='Módulo')
    ws_r.cell(row=7, column=2, value='Divergências (erros)')
    _header_row(ws_r, 7, 2)

    tipos = [('APONTAMENTOS','Apontamentos'), ('EMPRESTIMOS','Empréstimos'), ('COPARTICIPACAO','Coparticipação')]
    cores = ['C62828', '1565C0', '2E7D32']
    for i, ((tipo, label), cor) in enumerate(zip(tipos, cores), 8):
        n = sum(1 for d in divergencias if d.tipo == tipo)
        erros = sum(1 for d in divergencias if d.tipo == tipo and d.gravidade == 'ERRO')
        ws_r.cell(row=i, column=1, value=label).font = Font(name='Arial', size=10)
        txt = f'{n} ({erros} erros)' if n > 0 else '✅ OK — sem divergências'
        c = ws_r.cell(row=i, column=2, value=txt)
        c.font = Font(name='Arial', size=10, color=cor if n > 0 else '1A7F37', bold=n > 0)

    total = len(divergencias)
    erros_t = sum(1 for d in divergencias if d.gravidade == 'ERRO')
    status = '✅ SEM DIVERGÊNCIAS' if total == 0 else f'⚠️ {total} DIVERGÊNCIAS — {erros_t} ERROS'
    ws_r.cell(row=12, column=1, value='Status Geral').font = Font(name='Arial', size=11, bold=True)
    c = ws_r.cell(row=12, column=2, value=status)
    c.font = Font(name='Arial', size=11, bold=True, color='1A7F37' if total == 0 else 'C62828')

    for tipo, nome_aba in tipos:
        _aba_divs(wb.create_sheet(nome_aba), [d for d in divergencias if d.tipo == tipo], nome_aba)

    wb.save(destino)


# ═══════════════════════════════════════════════════════════════════════════════
# Auto-detecção de arquivos na pasta
# ═══════════════════════════════════════════════════════════════════════════════

def _find(pasta: str, palavras: list) -> Optional[str]:
    """Procura arquivo na pasta cujo nome contenha alguma das palavras (case-insensitive)."""
    for f in os.listdir(pasta):
        fl = f.lower()
        if any(p.lower() in fl for p in palavras):
            return os.path.join(pasta, f)
    return None


def detectar_arquivos(pasta: str) -> dict:
    return {
        'extrato':      _find(pasta, ['apontamentos__-_sistema', 'aponatmentos__-_sistema',
                                       'extrato', 'sistema']),
        'apontamentos': _find(pasta, ['apontamentos_', 'apontamentos-']),
        'emprestimos':  _find(pasta, ['emprestimos', 'empréstimos']),
        'coop':         _find(pasta, ['coparticipacao', 'cooparticipacao',
                                       'descontos_de_coo', 'coo-participacao']),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════════════

SEP = '─' * 60

def _ok(msg):  print(f'  ✅  {msg}')
def _err(msg): print(f'  ❌  {msg}')
def _warn(msg):print(f'  ⚠️   {msg}')
def _info(msg):print(f'  ℹ️   {msg}')


def main():
    parser = argparse.ArgumentParser(
        description='Validação de Folha de Pagamento — Domínio vs Excels',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos:
  python validador.py --pasta ./06-2026
  python validador.py --extrato folha.pdf --apontamentos apo.xls --emprestimos emp.xlsx
  python validador.py --extrato folha.pdf --apontamentos apo.xls --emprestimos emp.xlsx --coop coop.xls
"""
    )
    parser.add_argument('--pasta',         help='Pasta com todos os arquivos (auto-detecção)')
    parser.add_argument('--extrato',       help='PDF extrato mensal (Domínio)')
    parser.add_argument('--apontamentos',  help='Excel de apontamentos')
    parser.add_argument('--emprestimos',   help='Excel de empréstimos')
    parser.add_argument('--coop',          help='Excel de coparticipação (opcional)')
    parser.add_argument('--saida',         help='Caminho do relatório de saída (padrão: pasta de origem)')
    args = parser.parse_args()

    print()
    print('╔══════════════════════════════════════════════════════════╗')
    print('║       VALIDAÇÃO DE FOLHA DE PAGAMENTO — Huty             ║')
    print('╚══════════════════════════════════════════════════════════╝')
    print()

    # ── Resolve caminhos ──────────────────────────────────────────────────────
    if args.pasta:
        if not os.path.isdir(args.pasta):
            _err(f'Pasta não encontrada: {args.pasta}'); sys.exit(1)
        print(f'  Pasta: {os.path.abspath(args.pasta)}')
        detectados = detectar_arquivos(args.pasta)
        extrato_path     = args.extrato      or detectados['extrato']
        apontamentos_path = args.apontamentos or detectados['apontamentos']
        emprestimos_path  = args.emprestimos  or detectados['emprestimos']
        coop_path         = args.coop         or detectados['coop']
        pasta_saida       = args.pasta
    else:
        extrato_path      = args.extrato
        apontamentos_path = args.apontamentos
        emprestimos_path  = args.emprestimos
        coop_path         = args.coop
        pasta_saida       = os.path.dirname(os.path.abspath(extrato_path)) if extrato_path else '.'

    # ── Valida obrigatórios ───────────────────────────────────────────────────
    faltando = []
    if not extrato_path or not os.path.exists(extrato_path):
        faltando.append('Extrato PDF (Domínio)')
    if not apontamentos_path or not os.path.exists(apontamentos_path):
        faltando.append('Apontamentos Excel')
    if not emprestimos_path or not os.path.exists(emprestimos_path):
        faltando.append('Empréstimos Excel')

    if faltando:
        _err('Arquivos obrigatórios não encontrados:')
        for f in faltando: print(f'       • {f}')
        print()
        print('  Use --pasta <caminho> ou informe os arquivos individualmente.')
        print('  Execute  python validador.py --help  para mais detalhes.')
        sys.exit(1)

    # ── Leitura ───────────────────────────────────────────────────────────────
    print(SEP)
    print('  LENDO ARQUIVOS')
    print(SEP)

    print(f'  PDF extrato:      {os.path.basename(extrato_path)}')
    folha, empresa, competencia = parse_extrato_pdf(extrato_path)
    _ok(f'{len(folha)} funcionários lidos  |  Empresa: {empresa}  |  Comp: {competencia}')

    print(f'  Apontamentos:     {os.path.basename(apontamentos_path)}')
    apo = parse_apontamentos(apontamentos_path)
    _ok(f'{len(apo)} linhas de apontamentos')

    print(f'  Empréstimos:      {os.path.basename(emprestimos_path)}')
    emp = parse_emprestimos(emprestimos_path)
    _ok(f'{len(emp)} contratos')

    coop = None
    if coop_path and os.path.exists(coop_path):
        print(f'  Coparticipação:   {os.path.basename(coop_path)}')
        coop = parse_coop_excel(coop_path)
        _ok(f'{len(coop)} funcionários')
    else:
        _info('Coparticipação não informada — módulo ignorado')

    # ── Validação ─────────────────────────────────────────────────────────────
    print()
    print(SEP)
    print('  VALIDANDO')
    print(SEP)

    divs = []
    d_apo = validate_apontamentos(apo, folha)
    divs += d_apo
    erros_apo = sum(1 for d in d_apo if d.gravidade == 'ERRO')
    if d_apo:
        _warn(f'Apontamentos:   {len(d_apo)} divergências  ({erros_apo} erros)')
    else:
        _ok('Apontamentos:   sem divergências')

    d_emp = validate_emprestimos(emp, folha)
    divs += d_emp
    erros_emp = sum(1 for d in d_emp if d.gravidade == 'ERRO')
    if d_emp:
        _warn(f'Empréstimos:    {len(d_emp)} divergências  ({erros_emp} erros)')
        for d in d_emp:
            print(f'       [{d.gravidade}] {d.nome[:35]:35s} | {d.descricao}')
            print(f'              Excel=R${d.valor_excel:.2f}  Folha=R${d.valor_folha:.2f}  Diff={d.diff:+.2f}')
    else:
        _ok('Empréstimos:    sem divergências')

    if coop is not None:
        d_coop = validate_coop(coop, folha)
        divs += d_coop
        erros_coop = sum(1 for d in d_coop if d.gravidade == 'ERRO')
        if d_coop:
            _warn(f'Coparticipação: {len(d_coop)} divergências  ({erros_coop} erros)')
        else:
            _ok('Coparticipação: sem divergências')

    # ── Relatório ─────────────────────────────────────────────────────────────
    print()
    print(SEP)
    print('  GERANDO RELATÓRIO')
    print(SEP)

    comp_safe = competencia.replace('/', '-')
    nome_rel = args.saida or os.path.join(pasta_saida, f'Validacao_Folha_{comp_safe}.xlsx')
    gerar_relatorio(divs, competencia, empresa, len(folha), nome_rel)

    total = len(divs)
    erros_t = sum(1 for d in divs if d.gravidade == 'ERRO')
    print()
    if total == 0:
        _ok('RESULTADO FINAL: SEM DIVERGÊNCIAS ✅')
    else:
        _warn(f'RESULTADO FINAL: {total} divergências encontradas  ({erros_t} erros)')
    print()
    _ok(f'Relatório salvo em: {os.path.abspath(nome_rel)}')
    print()


if __name__ == '__main__':
    main()
