"""
validador_v2.py — Gera relatório completo no formato profissional
11 abas: Resumo Executivo, Comparativo Inputs, Divergências, Pontos de Atenção,
         Empréstimos, Resumo Funcionários, Folha Rubricas,
         Input Apontamentos, Input Coparticipação, Base Empréstimos, Fonte e Controle
"""

import re, io, math, datetime, os
from dataclasses import dataclass, field
from typing import Optional
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════
# CORES E ESTILOS
# ═══════════════════════════════════════════════════════════════════
C_AZUL      = '1F3864'
C_AZUL_LT   = 'D6E4F7'
C_VERDE     = '1A7A42'
C_VERDE_LT  = 'E2EFDA'
C_VERMELHO  = 'C62828'
C_VERM_LT   = 'FFDCE0'
C_AMARELO   = 'B45309'
C_AMAR_LT   = 'FFF3CD'
C_CINZA     = 'F2F2F2'
C_CINZA_ESC = '595959'
C_BRANCO    = 'FFFFFF'

def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def _font(bold=False, color='000000', size=10, name='Calibri'):
    return Font(bold=bold, color=color, size=size, name=name)

def _border(color='BFBFBF'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _header_row(ws, row, cols, bg=C_AZUL, fg=C_BRANCO, height=22):
    ws.row_dimensions[row].height = height
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _fill(bg)
        cell.font = _font(bold=True, color=fg, size=10)
        cell.alignment = _align('center', 'center', wrap=True)
        cell.border = _border()

def _title_row(ws, row, text, cols, bg=C_AZUL):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.fill = _fill(bg)
    cell.font = _font(bold=True, color=C_BRANCO, size=12)
    cell.alignment = _align('center', 'center')
    ws.row_dimensions[row].height = 26

def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _write_row(ws, row, values, bg=None, bold=False, num_cols=None):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c)
        cell.value = v
        cell.font = _font(bold=bold, size=9)
        cell.alignment = _align('left', 'center', wrap=True)
        cell.border = _border()
        if bg:
            cell.fill = _fill(bg)
        # Formata números
        if isinstance(v, float):
            cell.number_format = '#,##0.00'
            cell.alignment = _align('right', 'center')
    ws.row_dimensions[row].height = 15

# ═══════════════════════════════════════════════════════════════════
# ESTRUTURAS
# ═══════════════════════════════════════════════════════════════════
@dataclass
class FuncionarioFolha:
    matricula: int
    nome: str
    salario: float = 0.0
    horas_mes: float = 220.0
    rubricas: dict = field(default_factory=dict)        # {cod: valor}
    rubricas_raw: list = field(default_factory=list)    # [{cod,val,tipo,desc}]
    emprestimos: list = field(default_factory=list)
    proventos: float = 0.0
    descontos: float = 0.0
    informativas: float = 0.0
    liquido: float = 0.0

TOLERANCE = 0.10

# ═══════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════
def _to_float(v) -> float:
    if v is None or (isinstance(v, float) and math.isnan(v)): return 0.0
    if isinstance(v, str):
        v = v.replace('R$','').strip()
        if re.search(r'\d\.\d{3},', v): v = v.replace('.','').replace(',','.')
        else: v = v.replace(',','.')
        v = re.sub(r'[^\d.]', '', v)
        try: return float(v)
        except: return 0.0
    try: return float(v)
    except: return 0.0

def _is_cancelado(v): return isinstance(v, str) and 'CANCEL' in v.upper()

def _parse_rubrica_line(line):
    results = []
    segments = re.split(r'(?<=[ PD\*])\s+(?=\d{1,5}[^\d,.])', line)
    for seg in segments:
        seg = seg.strip()
        mc = re.match(r'^(\d{1,5})', seg)
        if not mc: continue
        cod = mc.group(1)
        mv = re.search(r'([\d]{1,3}(?:\.\d{3})*(?:,\d+)?)\s*([PD\*])\s*$', seg)
        if not mv: continue
        val = _to_float(mv.group(1))
        ind = mv.group(2)
        if val > 0:
            # Extrai descrição
            desc_match = re.match(r'^\d{1,5}\s+(.+?)\s+[\d.,]+\s+[\d.,]+\s*[PD\*]', seg)
            desc = desc_match.group(1).strip() if desc_match else ''
            results.append((cod, val, ind, desc))
    return results

# ═══════════════════════════════════════════════════════════════════
# PARSE PDF
# ═══════════════════════════════════════════════════════════════════
_EH = re.compile(r'Empr\.\:\s*(\d+)\s*(.+?)\s+Situa[çc]')
_CH = re.compile(r'Contr\:\s*(\d+)\s*(.+?)\s+Situa[çc]')
_SL = re.compile(r'Sal[aá]rio:\s*([\d.,]+)')
_HL = re.compile(r'Horas\s+M[eê]s:\s*([\d.,]+)')
_LL = re.compile(r'^(\d{3,4})\s+DESC[\.\s]+EMP[\.\s]+CRED[\.\s]+TRAB\s+N[oOº°]?\s*([\w]+)\s+([\d.,]+)\s+([\d.,]+)\s*D\s*$', re.I)
_EL = re.compile(r'Empresa:\s*\d+\s*-\s*(.+?)(?:\s+P[áa]gina|\s+CNPJ|$)')
_CL = re.compile(r'Compet[êe]ncia:\s*(\d{2}/\d{4})')
_PL = re.compile(r'Proventos:\s*([\d.,]+)')
_DL = re.compile(r'Descontos:\s*([\d.,]+)')
_LL2 = re.compile(r'Líquido:\s*Informativa:\s*([\d.,]+).*?([\d.,]+)\s*$')

def parse_extrato_pdf(path):
    funcs = {}; cur = None; empresa = competencia = ''
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text(x_tolerance=3, y_tolerance=3) or ''
            for line in text.splitlines():
                line = line.strip()
                if not line: continue
                if not empresa:
                    m = _EL.search(line)
                    if m: empresa = m.group(1).strip()
                if not competencia:
                    m = _CL.search(line)
                    if m: competencia = m.group(1)
                # Para ao encontrar resumo geral (evita acumular totais nos últimos funcionários)
                if 'Resumo por Rubrica' in line or 'Total Geral Proventos' in line or 'mo das Bases' in line:
                    if cur:
                        funcs[cur.matricula] = cur
                        cur = None
                    break

                m = _EH.search(line) or _CH.search(line)
                if m:
                    if cur: funcs[cur.matricula] = cur
                    nome = re.split(r'\s+(?:CPF|Situa)', m.group(2))[0].strip()
                    cur = FuncionarioFolha(matricula=int(m.group(1)), nome=nome)
                    continue
                if cur is None: continue
                ms = _SL.search(line)
                if ms and cur.salario == 0.0: cur.salario = _to_float(ms.group(1))
                mh = _HL.search(line)
                if mh:
                    v = _to_float(mh.group(1))
                    if v > 0: cur.horas_mes = v
                # Totais
                mp = _PL.search(line)
                if mp and 'Proventos:' in line: cur.proventos = _to_float(mp.group(1))
                md = _DL.search(line)
                if md and 'Descontos:' in line: cur.descontos = _to_float(md.group(1))
                ml2 = _LL2.search(line)
                if ml2 and 'Líquido:' in line:
                    cur.informativas = _to_float(ml2.group(1))
                    cur.liquido = _to_float(ml2.group(2))
                # Empréstimos
                ml = _LL.match(line)
                if ml:
                    valor = _to_float(ml.group(4))
                    if valor > 0:
                        cur.emprestimos.append({'contrato': ml.group(2).strip(), 'valor': valor, 'rubrica': ml.group(1)})
                    continue
                # Rubricas
                for cod, val, ind, desc in _parse_rubrica_line(line):
                    # Rubricas informativas (*): não acumula, usa último valor
                    # Rubricas P/D: acumula (para casos de dependentes como odonto)
                    if ind == '*':
                        if cod not in cur.rubricas:
                            cur.rubricas[cod] = val
                            cur.rubricas_raw.append({'cod': cod, 'val': val, 'tipo': ind, 'desc': desc})
                        # se já existe, ignora (evita duplicação de informativas entre funcionários da mesma página)
                    else:
                        cur.rubricas[cod] = cur.rubricas.get(cod, 0.0) + val
                        cur.rubricas_raw.append({'cod': cod, 'val': val, 'tipo': ind, 'desc': desc})
    if cur: funcs[cur.matricula] = cur
    return funcs, empresa, competencia

# ═══════════════════════════════════════════════════════════════════
# PARSE EXCELS
# ═══════════════════════════════════════════════════════════════════
def parse_apontamentos(path):
    ext = path.lower().split('.')[-1]
    engine = 'xlrd' if ext == 'xls' else 'openpyxl'
    raw = pd.read_excel(path, header=None, engine=engine)
    header_row = code_row = None
    for i, row in raw.iterrows():
        vals = [str(v) for v in row.values]
        if any('digo Empregado' in v or 'Codigo Empregado' in v for v in vals):
            header_row = i
        if header_row is not None and i == header_row + 1:
            code_row = i; break
    if header_row is None: raise ValueError(f'Apontamentos: cabeçalho não encontrado.')
    col_codes = {}
    if code_row is not None:
        for ci, val in enumerate(raw.iloc[code_row]):
            s = str(val).strip().split('.')[0]
            if s.isdigit(): col_codes[ci] = s
    # Nomes das rubricas (linha do cabeçalho)
    col_names = {}
    for ci, val in enumerate(raw.iloc[header_row]):
        s = str(val).strip()
        if s and s != 'nan': col_names[ci] = s
    result = {}
    raw_rows = []
    for idx in range(header_row + 2, len(raw)):
        row = raw.iloc[idx]
        mat_raw = row.iloc[1]
        if pd.isna(mat_raw): continue
        try: mat = int(float(str(mat_raw).strip()))
        except: continue
        if mat <= 0: continue
        nome_raw = str(row.iloc[2]) if len(row) > 2 else ''
        data = {}
        for ci, rub in col_codes.items():
            v = row.iloc[ci]
            is_canc = _is_cancelado(v)
            val = 0.0 if is_canc else _to_float(v)
            status = 'CANCELADO' if is_canc else ('CANCELAR DESCONTO' if isinstance(v, str) and 'CANCELAR' in v.upper() else ('VALOR' if val > 0 else None))
            data[rub] = {'val': val, 'status': status, 'raw': str(v) if v is not None else ''}
            if status:
                raw_rows.append({
                    'origem': 'APONTAMENTOS', 'cod_emp': mat, 'nome_input': nome_raw,
                    'rubrica': rub, 'status_input': status, 'valor_esperado': val if val > 0 else None,
                    'obs_input': f'Input indica {status}; rubrica deve estar ausente ou zerada.' if 'CANCEL' in status else None
                })
        result[mat] = {'data': data, 'nome': nome_raw}
    return result, raw_rows

def parse_emprestimos(path):
    ext = path.lower().split('.')[-1]
    engine = 'xlrd' if ext == 'xls' else 'openpyxl'
    df = pd.read_excel(path, engine=engine)
    df.columns = [str(c).strip() for c in df.columns]
    records = []
    for _, row in df.iterrows():
        c = str(row.get('contrato', '')).strip()
        if not c or c.lower() in ('nan','none',''): continue
        p = _to_float(row.get('valorParcela'))
        if p <= 0: continue
        records.append({
            'nome': str(row.get('nomeTrabalhador','')).strip(),
            'contrato': c, 'valorParcela': p,
            **{k: row.get(k) for k in df.columns if k not in ['nomeTrabalhador','contrato','valorParcela']}
        })
    return records

def parse_coop_excel(path):
    ext = path.lower().split('.')[-1]
    engine = 'xlrd' if ext == 'xls' else 'openpyxl'
    raw = pd.read_excel(path, header=None, engine=engine)
    header_row = None
    for i, row in raw.iterrows():
        vals = [str(v).strip() for v in row.values]
        if any(v in ('Cód.','Cod.','Código') for v in vals): header_row = i; break
    if header_row is None: raise ValueError('Coparticipação: cabeçalho não reconhecido.')
    header = raw.iloc[header_row].tolist()
    col_cod = col_227 = col_203 = col_nome = None
    for i, h in enumerate(header):
        h = str(h).strip()
        if h in ('Cód.','Cod.','Código'): col_cod = i
        elif '227' in h: col_227 = i
        elif '203' in h: col_203 = i
        elif 'nome' in h.lower() or 'func' in h.lower(): col_nome = i
    if col_cod is None: col_cod = 1
    if col_227 is None: col_227 = 3
    if col_203 is None: col_203 = 6

    result = {}; raw_rows = []
    for idx in range(header_row + 1, len(raw)):
        row = raw.iloc[idx]
        mat_raw = row.iloc[col_cod]
        if pd.isna(mat_raw): continue
        try: mat = int(float(str(mat_raw).strip()))
        except: continue
        if mat <= 0: continue
        nome = str(row.iloc[col_nome]) if col_nome else ''
        v227 = _to_float(row.iloc[col_227])
        v203 = _to_float(row.iloc[col_203]) if col_203 < len(row) else 0.0
        if mat in result:
            result[mat]['227'] += v227; result[mat]['203'] += v203
        else:
            result[mat] = {'227': v227, '203': v203, 'nome': nome}
        if v227 > 0:
            raw_rows.append({'origem':'COOPARTICIPACAO','cod_emp':mat,'nome_input':nome,'rubrica':'227',
                             'descricao_input':'Plano de saúde informativa','valor_esperado':v227,'status_input':'VALOR'})
        if v203 > 0:
            raw_rows.append({'origem':'COOPARTICIPACAO','cod_emp':mat,'nome_input':nome,'rubrica':'203',
                             'descricao_input':'Fator moderador Unimed','valor_esperado':v203,'status_input':'VALOR'})
    return result, raw_rows

# ═══════════════════════════════════════════════════════════════════
# DESCRIÇÕES DE RUBRICAS
# ═══════════════════════════════════════════════════════════════════
DESCRICOES = {
    '227': 'Plano de saúde informativa', '230': 'VR Local Informativa',
    '231': 'VR Cartão Informativa', '224': 'VT Cartão Informativa',
    '8111': 'Desconto Plano de Saúde', '202': 'Assistência Odontológica',
    '203': 'Fator Moderador Unimed', '204': 'Desconto Vale Refeição',
    '48': 'Desc. Vale Transporte', '150': 'Horas Extras 50%', '200': 'Horas Extras 100%',
    '8069': 'Faltas', '8792': 'Faltas parciais', '8794': 'Faltas DSR',
}

# ═══════════════════════════════════════════════════════════════════
# ANÁLISE PRINCIPAL
# ═══════════════════════════════════════════════════════════════════
def analisar(funcs, apo_data, apo_raw, emp_list, coop_data, coop_raw):
    comparativo = []  # Todas as linhas comparadas
    divergencias = []
    pontos_atencao = []
    emp_resultado = []

    # ── Apontamentos ─────────────────────────────────────────────
    RUBRICAS_DIRETAS = {
        '227':'Plano de saúde informativa','230':'VR Local Informativa',
        '231':'VR Cartão Informativa','224':'VT Cartão Informativa',
        '8111':'Desconto Plano de Saúde','202':'Assistência Odontológica',
        '203':'Fator Moderador Unimed','204':'Desconto Vale Refeição','48':'Desc. Vale Transporte'
    }
    RUBRICAS_HE = {'150':('Horas Extras 50%',1.5),'200':('Horas Extras 100%',2.0)}

    for mat, apo in apo_data.items():
        func = funcs.get(mat)
        nome_input = apo.get('nome','')
        data = apo.get('data', {})

        if func is None:
            pontos_atencao.append({
                'origem':'APONTAMENTOS','cod_emp':mat,'nome_input':nome_input,
                'rubrica':'—','descricao_input':'Matrícula não localizada na folha',
                'valor_esperado':None,'status_input':'ATENÇÃO',
                'obs_input':'Matrícula presente no Excel de apontamentos mas não encontrada no PDF da folha.',
                'valor_folha':None,'diferenca':None,'status':'ATENÇÃO',
                'critica':'Verificar se funcionário foi incluído no processamento.'
            })
            continue

        hm = func.horas_mes if func.horas_mes > 0 else 220.0
        sh = func.salario / hm if func.salario > 0 else 0.0

        for cod, desc in RUBRICAS_DIRETAS.items():
            info = data.get(cod, {})
            vx = info.get('val', 0.0) if info else 0.0
            status_input = info.get('status') if info else None
            vf = func.rubricas.get(cod, 0.0)
            diff = round(vx - vf, 2)

            if status_input in ('CANCELADO', 'CANCELAR DESCONTO'):
                obs = f'Input indica {status_input}; rubrica deve estar ausente ou zerada.'
                if vf > 0:
                    pontos_atencao.append({
                        'origem':'APONTAMENTOS','cod_emp':mat,'nome_input':func.nome,
                        'rubrica':cod,'descricao_input':desc,'valor_esperado':None,
                        'status_input':status_input,'obs_input':obs,
                        'valor_folha':vf,'diferenca':None,'status':'ATENÇÃO',
                        'critica':f'Input cancelado mas folha tem lançamento R${vf:.2f}. Verificar.'
                    })
                else:
                    row = {'origem':'APONTAMENTOS','cod_emp':mat,'nome_input':func.nome,
                           'rubrica':cod,'descricao_input':desc,'valor_esperado':None,
                           'referencia_esperada':None,'status_input':status_input,
                           'obs_input':obs,'valor_folha':None,'referencia_folha':None,
                           'diferenca':0,'status':'OK','critica':'Coerente: input cancelado e folha sem lançamento.'}
                    comparativo.append(row)
                continue

            # NaN no Excel = não informado = ponto de atenção, não divergência
            info = data.get(cod, {})
            excel_was_filled = bool(info and info.get('status') is not None)

            if not excel_was_filled:
                if vf > 0:
                    raw_entry = next((r for r in func.rubricas_raw if r['cod'] == cod), {})
                    pontos_atencao.append({
                        'origem': 'APONTAMENTOS', 'cod_emp': mat, 'nome_input': None,
                        'rubrica': cod, 'descricao_input': None,
                        'valor_esperado': None, 'referencia_esperada': None,
                        'status_input': None, 'obs_input': None,
                        'valor_folha': vf, 'referencia_folha': raw_entry.get('val'),
                        'diferenca': None, 'status': 'ATENÇÃO',
                        'critica': 'Lancamento existente na folha sem input numerico nas planilhas auxiliares.',
                        'nome': func.nome, 'descricao_folha': raw_entry.get('desc', '')
                    })
                continue

            if vx == 0 and vf == 0:
                continue

            diff = round(vx - vf, 2)
            row = {
                'origem': 'APONTAMENTOS', 'cod_emp': mat, 'nome_input': func.nome,
                'rubrica': cod, 'descricao_input': desc,
                'valor_esperado': vx if vx > 0 else None,
                'referencia_esperada': None,
                'status_input': 'VALOR' if vx > 0 else None,
                'obs_input': None, 'valor_folha': vf, 'referencia_folha': 0,
                'diferenca': diff,
                'status': 'OK' if abs(diff) <= TOLERANCE else 'DIVERGENCIA',
                'critica': 'Coerente.' if abs(diff) <= TOLERANCE else f'Valor diferente. Excel={vx:.2f} Folha={vf:.2f}'
            }
            if abs(diff) > TOLERANCE:
                divergencias.append(dict(row))
            comparativo.append(row)

        # HE — só processa se Excel tem valor explícito (não nan)
        for cod, (desc, mult) in RUBRICAS_HE.items():
            info = data.get(cod, {})
            excel_filled = bool(info and info.get('status') is not None)
            horas = info.get('val', 0.0) if info else 0.0
            vf = func.rubricas.get(cod, 0.0)

            if not excel_filled:
                # HE não informado no Excel - ponto de atenção se houver na folha
                if vf > 0:
                    pontos_atencao.append({
                        'origem':'APONTAMENTOS','cod_emp':mat,'nome_input':None,
                        'rubrica':cod,'descricao_input':None,
                        'valor_esperado':None,'referencia_esperada':None,
                        'status_input':None,'obs_input':None,
                        'valor_folha':vf,'referencia_folha':None,
                        'diferenca':None,'status':'ATENÇÃO',
                        'critica':f'HE na folha (R${vf:.2f}) sem registro no Excel de apontamentos.',
                        'nome':func.nome,'descricao_folha':desc
                    })
                continue

            if horas == 0 and vf == 0: continue
            vx = round(horas * sh * mult, 2) if sh > 0 else 0.0
            diff = round(vx - vf, 2)
            row = {
                'origem':'APONTAMENTOS','cod_emp':mat,'nome_input':func.nome,
                'rubrica':cod,'descricao_input':f'{desc} ({horas}h × R${sh:.4f} × {mult})',
                'valor_esperado':vx,'referencia_esperada':horas,
                'status_input':'VALOR','obs_input':None,
                'valor_folha':vf,'referencia_folha':0,
                'diferenca':diff,
                'status':'OK' if abs(diff) <= TOLERANCE else 'DIVERGÊNCIA',
                'critica':'Coerente.' if abs(diff) <= TOLERANCE else 'Valor calculado difere da folha.'
            }
            if abs(diff) > TOLERANCE: divergencias.append(dict(row))
            comparativo.append(row)

    # ── Coparticipação ────────────────────────────────────────────
    for mat, coop in coop_data.items():
        func = funcs.get(mat)
        nome_input = coop.get('nome','')
        for cod, desc in [('227','Plano de saúde informativa'),('203','Fator Moderador Unimed')]:
            vx = coop.get(cod, 0.0)
            vf = func.rubricas.get(cod, 0.0) if func else 0.0
            if vx == 0 and vf == 0: continue
            diff = round(vx - vf, 2)
            nome = func.nome if func else nome_input
            row = {
                'origem':'COOPARTICIPACAO','cod_emp':mat,'nome_input':nome,
                'rubrica':cod,'descricao_input':desc,'valor_esperado':vx if vx>0 else None,
                'referencia_esperada':None,'status_input':'VALOR' if vx>0 else None,
                'obs_input':None,'valor_folha':vf if vf>0 else None,'referencia_folha':None,
                'diferenca':diff,
                'status':'OK' if abs(diff) <= TOLERANCE else 'DIVERGÊNCIA',
                'critica':'Coerente.' if abs(diff) <= TOLERANCE else f'Rubrica informada no input não localizada na folha.' if vf==0 else f'Valor diferente do input.'
            }
            if abs(diff) > TOLERANCE: divergencias.append(dict(row))
            comparativo.append(row)

    # ── Empréstimos ───────────────────────────────────────────────
    # Agrupa por funcionário (nome normalizado)
    def norm_nome(n):
        return re.sub(r'\s+', ' ', n.upper().strip())

    # Folha: totais por funcionário
    folha_emp_por_func = {}
    for func in funcs.values():
        if func.emprestimos:
            n = norm_nome(func.nome)
            if n not in folha_emp_por_func:
                folha_emp_por_func[n] = {'nome':func.nome,'total':0,'contratos':[],'rubricas':[]}
            for loan in func.emprestimos:
                folha_emp_por_func[n]['total'] += loan['valor']
                folha_emp_por_func[n]['contratos'].append(loan['contrato'])
                folha_emp_por_func[n]['rubricas'].append(loan['rubrica'])

    # Excel: totais por funcionário
    excel_emp_por_func = {}
    for rec in emp_list:
        n = norm_nome(rec['nome'])
        if n not in excel_emp_por_func:
            excel_emp_por_func[n] = {'nome':rec['nome'],'total':0,'contratos':[]}
        excel_emp_por_func[n]['total'] += rec['valorParcela']
        excel_emp_por_func[n]['contratos'].append(rec['contrato'])

    todos_nomes = set(list(folha_emp_por_func.keys()) + list(excel_emp_por_func.keys()))
    for n in sorted(todos_nomes):
        folha = folha_emp_por_func.get(n, {'nome':n,'total':0,'contratos':[],'rubricas':[]})
        excel = excel_emp_por_func.get(n, {'nome':n,'total':0,'contratos':[]})
        diff = round(excel['total'] - folha['total'], 2)
        status = 'OK' if abs(diff) <= TOLERANCE else 'DIVERGÊNCIA'
        critica = 'Coerente.' if status == 'OK' else 'Total de empréstimos na folha difere da planilha EMPRESTIMOS; conferir contrato(s).'
        emp_resultado.append({
            'nome_norm': n,
            'nome_input': excel.get('nome', n),
            'valor_input': excel['total'] if excel['total']>0 else None,
            'contratos_input': ', '.join(excel['contratos']) if excel['contratos'] else None,
            'nome_folha': folha.get('nome', n),
            'valor_folha': folha['total'] if folha['total']>0 else None,
            'rubricas': ', '.join(folha['rubricas']) if folha.get('rubricas') else None,
            'diferenca': diff if abs(diff) > TOLERANCE else 0,
            'status': status,
            'critica': critica
        })

    # Pontos de atencao por lancamento na folha sem input tratados no loop acima

    return comparativo, divergencias, pontos_atencao, emp_resultado

# ═══════════════════════════════════════════════════════════════════
# GERA RELATÓRIO EXCEL
# ═══════════════════════════════════════════════════════════════════
def gerar_relatorio(funcs, apo_data, apo_raw, emp_list, coop_data, coop_raw,
                    comparativo, divergencias, pontos_atencao, emp_resultado,
                    empresa, competencia, destino):

    wb = Workbook()
    emissao = datetime.datetime.now().strftime('%d/%m/%Y %H:%M')
    total_itens = len(comparativo)
    total_ok    = sum(1 for r in comparativo if r['status']=='OK')
    total_div   = len(divergencias)

    # ── 1. RESUMO EXECUTIVO ──────────────────────────────────────
    ws = wb.active; ws.title = 'Resumo Executivo'
    _set_col_widths(ws, [28, 30, 10, 10, 10, 10, 10, 10, 10, 10, 10])
    _title_row(ws, 1, f'Resumo Executivo — Análise Folha {empresa} {competencia}', 11)

    indicadores = [
        ('Empresa', empresa),
        ('Competência', competencia),
        ('Empregados na folha', len(funcs)),
        ('Total de proventos', sum(f.proventos for f in funcs.values())),
        ('Total de descontos', sum(f.descontos for f in funcs.values())),
        ('Líquido geral', sum(f.liquido for f in funcs.values())),
        ('Itens comparados contra inputs', total_itens),
        ('Itens OK', total_ok),
        ('Divergências encontradas', total_div),
        ('Pontos de atenção', len(pontos_atencao)),
        ('Taxa de conformidade', f'{(total_ok/total_itens*100):.1f}%' if total_itens>0 else 'N/A'),
        ('Emissão', emissao),
    ]
    ws.cell(row=3, column=1, value='Indicador').font = _font(bold=True, size=10)
    ws.cell(row=3, column=2, value='Resultado').font = _font(bold=True, size=10)
    _header_row(ws, 3, 2, bg=C_AZUL)

    for i, (k, v) in enumerate(indicadores, 4):
        ws.cell(row=i, column=1, value=k).font = _font(bold=True, size=10)
        c = ws.cell(row=i, column=2, value=v)
        c.font = _font(size=10)
        if isinstance(v, float): c.number_format = '#,##0.00'
        bg = C_VERDE_LT if k == 'Divergências encontradas' and v == 0 else (C_VERM_LT if k == 'Divergências encontradas' and v > 0 else C_CINZA)
        ws.cell(row=i, column=1).fill = _fill(C_CINZA)
        ws.cell(row=i, column=2).fill = _fill(bg)
        for c2 in range(1, 3):
            ws.cell(row=i, column=c2).border = _border()
        ws.row_dimensions[i].height = 16

    # Destaque status geral
    status_row = len(indicadores) + 5
    ws.merge_cells(start_row=status_row, start_column=1, end_row=status_row, end_column=2)
    cell = ws.cell(row=status_row, column=1)
    if total_div == 0:
        cell.value = '✅  SEM DIVERGÊNCIAS — Folha está em conformidade com os inputs'
        cell.fill = _fill(C_VERDE_LT)
        cell.font = _font(bold=True, color=C_VERDE, size=11)
    else:
        cell.value = f'⚠️  {total_div} DIVERGÊNCIA(S) ENCONTRADA(S) — Revisar antes de fechar a folha'
        cell.fill = _fill(C_VERM_LT)
        cell.font = _font(bold=True, color=C_VERMELHO, size=11)
    cell.alignment = _align('center', 'center')
    ws.row_dimensions[status_row].height = 24

    # ── 2. COMPARATIVO INPUTS ────────────────────────────────────
    ws2 = wb.create_sheet('Comparativo Inputs')
    _set_col_widths(ws2, [16,8,35,9,30,14,14,14,42,12,14,11,13,50])
    _title_row(ws2, 1, 'Comparativo Inputs — Todos os itens verificados', 14, bg=C_AZUL)
    headers = ['origem','cod_emp','nome_input','rubrica','descricao_input',
               'valor_esperado','referencia_esperada','status_input','obs_input',
               'valor_folha','referencia_folha','diferenca','status','critica']
    for c, h in enumerate(headers, 1):
        ws2.cell(row=2, column=c, value=h)
    _header_row(ws2, 2, len(headers))

    for r, row in enumerate(comparativo, 3):
        vals = [row.get(h) for h in headers]
        bg = C_VERDE_LT if row['status']=='OK' else C_VERM_LT
        _write_row(ws2, r, vals, bg=bg)

    # ── 3. DIVERGÊNCIAS ──────────────────────────────────────────
    ws3 = wb.create_sheet('Divergências')
    _set_col_widths(ws3, [16,8,35,9,30,14,14,14,42,12,14,11,13,50])
    _title_row(ws3, 1, f'Divergências — {len(divergencias)} item(ns) com inconsistência', 14, bg=C_VERMELHO)
    for c, h in enumerate(headers, 1):
        ws3.cell(row=2, column=c, value=h)
    _header_row(ws3, 2, len(headers), bg=C_VERMELHO)

    for r, row in enumerate(divergencias, 3):
        vals = [row.get(h) for h in headers]
        _write_row(ws3, r, vals, bg=C_VERM_LT, bold=True)

    if not divergencias:
        ws3.cell(row=3, column=1, value='✅  Nenhuma divergência encontrada.')
        ws3.cell(row=3, column=1).font = _font(bold=True, color=C_VERDE, size=11)

    # ── 4. PONTOS DE ATENÇÃO ─────────────────────────────────────
    ws4 = wb.create_sheet('Pontos de Atenção')
    _set_col_widths(ws4, [16,8,35,9,30,14,14,14,42,12,14,11,13,60,30,30])
    _title_row(ws4, 1, f'Pontos de Atenção — {len(pontos_atencao)} item(ns) para revisão manual', 16, bg=C_AMARELO)
    hpa = headers + ['nome','descricao_folha']
    for c, h in enumerate(hpa, 1):
        ws4.cell(row=2, column=c, value=h)
    _header_row(ws4, 2, len(hpa), bg=C_AMARELO)

    for r, row in enumerate(pontos_atencao, 3):
        vals = [row.get(h) for h in hpa]
        _write_row(ws4, r, vals, bg=C_AMAR_LT)

    # ── 5. EMPRÉSTIMOS ───────────────────────────────────────────
    ws5 = wb.create_sheet('Empréstimos')
    _set_col_widths(ws5, [35,35,14,60,35,14,30,12,14,60])
    _title_row(ws5, 1, 'Empréstimos — Comparativo por funcionário', 10, bg=C_AZUL)
    hemp = ['nome_norm','nome_input','valor_input','contratos_input','nome_folha','valor_folha','rubricas','diferenca','status','critica']
    for c, h in enumerate(hemp, 1):
        ws5.cell(row=2, column=c, value=h)
    _header_row(ws5, 2, len(hemp))
    for r, row in enumerate(emp_resultado, 3):
        vals = [row.get(h) for h in hemp]
        bg = C_VERDE_LT if row['status']=='OK' else C_VERM_LT
        _write_row(ws5, r, vals, bg=bg)

    # ── 6. RESUMO FUNCIONÁRIOS ───────────────────────────────────
    ws6 = wb.create_sheet('Resumo Funcionários')
    _set_col_widths(ws6, [8,35,6,14,14,14,14,12,14])
    _title_row(ws6, 1, 'Resumo por Funcionário', 9)
    hrf = ['cod_emp','nome','page','proventos','descontos','informativa','liquido','qtd_alertas','status_geral']
    for c, h in enumerate(hrf, 1):
        ws6.cell(row=2, column=c, value=h)
    _header_row(ws6, 2, len(hrf))

    for r, (mat, func) in enumerate(sorted(funcs.items()), 3):
        alertas = sum(1 for p in pontos_atencao if p.get('cod_emp') == mat) + \
                  sum(1 for d in divergencias if d.get('cod_emp') == mat)
        status = 'OK' if alertas == 0 else f'{alertas} alerta(s)'
        vals = [mat, func.nome, None, func.proventos, func.descontos, func.informativas, func.liquido, alertas, status]
        bg = C_VERDE_LT if alertas == 0 else C_AMAR_LT
        _write_row(ws6, r, vals, bg=bg)

    # ── 7. FOLHA RUBRICAS ────────────────────────────────────────
    ws7 = wb.create_sheet('Folha Rubricas')
    _set_col_widths(ws7, [8,35,9,6,14,14,35])
    _title_row(ws7, 1, 'Todas as rubricas extraídas do PDF da folha', 7)
    hfr = ['cod_emp','nome','rubrica','tipo','valor_folha','referencia_folha','descricao_folha']
    for c, h in enumerate(hfr, 1):
        ws7.cell(row=2, column=c, value=h)
    _header_row(ws7, 2, len(hfr))

    r = 3
    for mat, func in sorted(funcs.items()):
        for raw in func.rubricas_raw:
            vals = [mat, func.nome, raw['cod'], raw['tipo'], raw['val'], raw['val'], raw['desc']]
            _write_row(ws7, r, vals)
            r += 1

    # ── 8. INPUT APONTAMENTOS ────────────────────────────────────
    ws8 = wb.create_sheet('Input Apontamentos')
    _set_col_widths(ws8, [16,8,35,9,30,14,14,14,42])
    _title_row(ws8, 1, 'Input Apontamentos — Dados brutos lidos do Excel', 9, bg=C_CINZA_ESC)
    hia = ['origem','cod_emp','nome_input','rubrica','descricao_input','valor_esperado','referencia_esperada','status_input','obs_input']
    for c, h in enumerate(hia, 1):
        ws8.cell(row=2, column=c, value=h)
    _header_row(ws8, 2, len(hia))
    for r2, row in enumerate(apo_raw, 3):
        vals = [row.get(h) for h in hia]
        _write_row(ws8, r2, vals)

    # ── 9. INPUT COPARTICIPAÇÃO ──────────────────────────────────
    ws9 = wb.create_sheet('Input Coparticipação')
    _set_col_widths(ws9, [16,8,35,9,30,14,14,14,42])
    _title_row(ws9, 1, 'Input Coparticipação — Dados brutos lidos do Excel', 9, bg=C_CINZA_ESC)
    for c, h in enumerate(hia, 1):
        ws9.cell(row=2, column=c, value=h)
    _header_row(ws9, 2, len(hia))
    for r2, row in enumerate(coop_raw, 3):
        vals = [row.get(h) for h in hia]
        _write_row(ws9, r2, vals)

    # ── 10. BASE EMPRÉSTIMOS ─────────────────────────────────────
    ws10 = wb.create_sheet('Base Empréstimos')
    _set_col_widths(ws10, [10,30,16,16,35,35,16,16,14,14,14,14,14,14,14,14,14,60,12,16,16,16,35])
    _title_row(ws10, 1, 'Base Empréstimos — Dados completos do Excel', 23, bg=C_CINZA_ESC)

    if emp_list:
        emp_df = pd.DataFrame(emp_list)
        cols_emp = list(emp_df.columns)
        for c, h in enumerate(cols_emp, 1):
            ws10.cell(row=2, column=c, value=h)
        _header_row(ws10, 2, len(cols_emp))
        for r2, (_, row_df) in enumerate(emp_df.iterrows(), 3):
            vals = list(row_df.values)
            _write_row(ws10, r2, vals)

    # ── 11. FONTE E CONTROLE ─────────────────────────────────────
    ws11 = wb.create_sheet('Fonte e Controle')
    _set_col_widths(ws11, [28,35,60])
    _title_row(ws11, 1, 'Fonte e Controle — Rastreabilidade dos arquivos utilizados', 3)
    hfc = ['Fonte','Arquivo','Uso no relatório']
    for c, h in enumerate(hfc, 1):
        ws11.cell(row=2, column=c, value=h)
    _header_row(ws11, 2, 3)
    fontes = [
        ('Arquivo folha analisado', os.path.basename(path_extrato) if 'path_extrato' in dir() else 'PDF extrato', f'Extrato Mensal {competencia} — emissão {emissao}'),
        ('Input apontamentos', os.path.basename(path_apo) if 'path_apo' in dir() else 'APONTAMENTOS.xls', 'Rubricas 150, 200, 8111, 202, 8069, 8792, 8794, 37, 204, 227, 224, 230, 231 e 48'),
        ('Input coparticipação', os.path.basename(path_coop) if 'path_coop' in dir() else 'COOPARTICIPACAO.xls', 'Rubricas 227 e 203'),
        ('Input empréstimos', os.path.basename(path_emp) if 'path_emp' in dir() else 'EMPRESTIMOS.xlsx', 'Comparação de totais de empréstimos por funcionário'),
        ('Observação', 'Planilhas .xls convertidas apenas para extração de dados', 'Validar manualmente instruções textuais como CANCELADO/CANCELAR DESCONTO'),
    ]
    for r2, (f, a, u) in enumerate(fontes, 3):
        _write_row(ws11, r2, [f, a, u], bg=C_CINZA)

    wb.save(destino)
    print(f'✅ Relatório salvo: {destino}')
    return destino

# ═══════════════════════════════════════════════════════════════════
# EXECUÇÃO
# ═══════════════════════════════════════════════════════════════════
def executar(path_extrato, path_apo, path_emp, path_coop=None, destino=None):
    print('⏳ Lendo extrato PDF...')
    funcs, empresa, competencia = parse_extrato_pdf(path_extrato)
    print(f'  ✅ {len(funcs)} funcionários | {empresa} | {competencia}')

    print('⏳ Lendo apontamentos...')
    apo_data, apo_raw = parse_apontamentos(path_apo)
    print(f'  ✅ {len(apo_data)} linhas')

    print('⏳ Lendo empréstimos...')
    emp_list = parse_emprestimos(path_emp)
    print(f'  ✅ {len(emp_list)} contratos')

    coop_data = {}; coop_raw = []
    if path_coop:
        print('⏳ Lendo coparticipação...')
        coop_data, coop_raw = parse_coop_excel(path_coop)
        print(f'  ✅ {len(coop_data)} funcionários')

    print('⏳ Analisando...')
    comparativo, divergencias, pontos_atencao, emp_resultado = analisar(
        funcs, apo_data, apo_raw, emp_list, coop_data, coop_raw)
    print(f'  ✅ {len(comparativo)} itens | {len(divergencias)} divergências | {len(pontos_atencao)} atenções')

    if not destino:
        comp_safe = competencia.replace('/','_')
        destino = f'Validacao_Folha_{comp_safe}.xlsx'

    gerar_relatorio(funcs, apo_data, apo_raw, emp_list, coop_data, coop_raw,
                    comparativo, divergencias, pontos_atencao, emp_resultado,
                    empresa, competencia, destino)
    return destino

if __name__ == '__main__':
    import sys, argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--extrato', required=True)
    parser.add_argument('--apontamentos', required=True)
    parser.add_argument('--emprestimos', required=True)
    parser.add_argument('--coop', default=None)
    parser.add_argument('--saida', default=None)
    args = parser.parse_args()
    executar(args.extrato, args.apontamentos, args.emprestimos, args.coop, args.saida)
